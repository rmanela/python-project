import json
from flask import Flask, request
import os
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib import pagesizes
app = Flask(__name__)
UPLOAD_FOLDER = './uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
def get_file_path(file_name):
    return os.path.abspath(file_name)
@app.route('/uploads', methods=['POST'])
def upload_file():
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    if file:
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        wb = load_workbook(filename=os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        print(len(wb.sheetnames))
        print(wb.sheetnames)
        obj1= {'path': get_file_path(file.filename),
          'number of sheets': len(wb.sheetnames)
          }
        return obj1

def read_numeric_values_from_excel(file_path, sheet_name, column_letters):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    print(sheet)
    column_values = []
    for column_letter in column_letters:
        column = sheet[column_letter]
        for cell in column:
            if cell.value is not None:
                column_values.append(cell.value)
    return column_values
@app.route('/report', methods=['POST'])
def report_file():
    sheets = request.json['sheetsList']
    doch = []
    for sh in range(len(sheets)):
        letters = read_numeric_values_from_excel(request.json['filePath'], sheets[sh]['name'], sheets[sh]['columnLetters'])
        sum = 0
        print(letters)
        numbers = list(filter(lambda x: int(x) != None, letters))
        print('numbers', numbers)
        if sheets[sh]['active'] == 'sum':
            for i in numbers:
                sum += i
        else:
            index = 0
            for i in numbers:
                sum += i
                index += 1
            sum /= index
        doch.append({"sheetName": sheets[sh]['name'], "active": sheets[sh]['active'], "answer": sum})
        report_pdf_file(doch)
    return doch

def report_pdf_file(doch):
    pdf_file = "pdf.pdf"
    width, height =(850, 700)


    c = canvas.Canvas(pdf_file, pagesize=(850, 700))
    c.setFont("Helvetica", 8)

    formatted_data = json.dumps(doch, indent=4)

    c.drawString(50, height - 50, "JSON Data:")
    c.drawString(50, height - 70, formatted_data)
    c.save()
    print(f'PDF דוח נוצר ונשמר ב- {pdf_file} ')





if __name__ == '__main__':
    app.run()

# for running...
# {
#     "filePath": "C:\\מלכי פרל\\test.xlsx",
#     "sheetsList": [
#         {
#             "name": "sheet1",
#             "active": "avarage",
#             "columnLetters": ["A", "B"]
#         },
#         {
#             "name": "sheet2",
#             "active": "sum",
#             "columnLetters": ["A", "C"]
#         }
#     ]
#
# }