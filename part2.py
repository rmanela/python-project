from flask import Flask, request
import os
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from part1 import report_pdf_file

app = Flask(__name__)
UPLOAD_FOLDER = './uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
# level A
def get_file_path(file_name):
    return os.path.abspath(file_name)
@app.route('/sheets_amount', methods=['POST'])
def upload_file():
    file = request.files['file']
    print(file)
    if file.filename == '':
        return 'No selected file'
    if file:
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        wb = load_workbook(filename=os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
        print(len(wb.sheetnames))
        print(wb.sheetnames)
        obj1 = {
            'path': get_file_path(file.filename),
            'number of sheets': len(wb.sheetnames)
        }
        return obj1

def sum_excel_values(file_path):
    wb = load_workbook(file_path)
    total_sum = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    total_sum += int(cell)
    return total_sum
@app.route('/field_sum', methods=['POST'])
def sum_of_values_field_in_excel():
    file_path = request.json['filePath']
    sum = sum_excel_values(file_path)
    return {'sum': sum}

def plot_excel_sheets_sum(file_path):
    wb = load_workbook(file_path)
    sums = []
    sheet_names = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_sum = 0
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    sheet_sum += int(cell)
        sums.append(sheet_sum)
        sheet_names.append(sheet)
    plt.bar(sheet_names, sums)
    plt.xlabel('Sheet')
    plt.ylabel('Sum')
    plt.title('Sum of each sheet in Excel file')
    plt.show()
    return {'sums': sums, 'sheet_names': sheet_names}
@app.route('/plot', methods=['POST'])
def plot_excel():
    file_path = request.json['filePath']
    print(file_path)
    plot_excel_sheets_sum(file_path)
    return {"successfully": 100}

def avarage_excel_values(file_path):
    wb = load_workbook(file_path)
    total_sum = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sum = 0
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    sum += int(cell)
        total_sum.append(sum)
    sum = 0
    for i in total_sum:
        sum += i
    return sum / len(total_sum)
@app.route('/average', methods=['POST'])
def average_of_sheets():
    file_path = request.json['filePath']
    avg = avarage_excel_values(file_path)
    return {'average': avg}


# level B
@app.route('/totalPdf', methods=['POST'])
def doch_pdf_total():
    file_path = request.json['filePath']
    file_name = file_path[12:21]
    print(file_path, file_name)
    names = plot_excel_sheets_sum(file_path)['sheet_names']
    print(names)
    sums = plot_excel_sheets_sum(file_path)['sums']
    print(sums)
    avg = avarage_excel_values(file_path)
    print(avg)
    obj_pdf = {
        'file_name': file_name,
        'sheets': [],
        'average': avg
    }
    print(obj_pdf)
    for i in range(len(names)):
        obj_sheet = {
            'name_sheet': names[i],
            'sum_sheet': sums[i],
        }
        obj_pdf['sheets'].append(obj_sheet)
    report_pdf_file(obj_pdf)
    return obj_pdf



if __name__ == '__main__':
    app.run()
